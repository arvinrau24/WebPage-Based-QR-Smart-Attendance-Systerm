import datetime
import io
import base64
import uuid
import os

from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.db import transaction

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authtoken.models import Token

import qrcode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import timedelta

from .models import Course, Session, QRToken, AttendanceRecord, StudentProfile
from .serializers import CourseSerializer, SessionSerializer, QRTokenSerializer, AttendanceRecordSerializer
from .timetable_parser import parse_timetable_image
from .student_list_parser import parse_student_list
from alerts.models import Alert, AlertSessionExcuse
from alerts.engine import check_and_trigger_alerts, excuse_session
from alerts.serializers import AlertExcuseSerializer

User = get_user_model()

# Semester: 14 teaching weeks in 15 calendar weeks (7 class + 1 holiday + 7 class)
SEMESTER_CLASS_WEEKS_PER_BLOCK = 7
SEMESTER_HOLIDAY_WEEKS = 1
SEMESTER_TOTAL_CLASS_WEEKS = SEMESTER_CLASS_WEEKS_PER_BLOCK * 2
SEMESTER_CALENDAR_WEEKS = SEMESTER_TOTAL_CLASS_WEEKS + SEMESTER_HOLIDAY_WEEKS


def semester_end_from_start(semester_start):
    return (
        semester_start
        + datetime.timedelta(weeks=SEMESTER_CALENDAR_WEEKS)
        - datetime.timedelta(days=1)
    )


def calendar_week_offset(class_week_index):
    if class_week_index < SEMESTER_CLASS_WEEKS_PER_BLOCK:
        return class_week_index
    return class_week_index + SEMESTER_HOLIDAY_WEEKS


def _parse_semester_start(value):
    if not value:
        return None, 'Semester start date is required (YYYY-MM-DD).'
    try:
        return datetime.date.fromisoformat(str(value).strip()), None
    except ValueError:
        return None, 'Invalid date format. Use YYYY-MM-DD.'


# --- COURSES ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def course_list(request):
    if request.method == 'GET':
        if request.user.role == 'lecturer':
            courses = Course.objects.filter(lecturer=request.user)
        else:
            courses = request.user.enrolled_courses.all()
        return Response(CourseSerializer(courses, many=True).data)

    if request.method == 'POST':
        serializer = CourseSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(lecturer=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# == STUDENT ENROLLMENT ==
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_student_manual(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    matric_number = request.data.get('matric_number', '').strip().upper()
    full_name = request.data.get('full_name', '').strip()
    phone = request.data.get('phone', '').strip()
    email = request.data.get('email', '').strip()

    if not matric_number or not full_name:
        return Response({'error': 'Matric number and full name are required'}, status=400)

    student, created = StudentProfile.objects.update_or_create(
        course=course,
        matric_number=matric_number,
        defaults={
            'full_name': full_name,
            'phone': phone,
            'email': email,
        }
    )

    return Response({
        'message': f'Student {matric_number} {"added" if created else "updated"} successfully',
        'student': {
            'id': student.id,
            'matric_number': student.matric_number,
            'full_name': student.full_name,
            'phone': student.phone or '',
            'email': student.email or '',
            'section': student.section or '',
        }
    }, status=201 if created else 200)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_enrolled_students(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    students = StudentProfile.objects.filter(course=course).order_by('section', 'full_name')
    student_list = [{
        'id': s.id,
        'matric_number': s.matric_number,
        'full_name': s.full_name,
        'phone': s.phone or '',
        'email': s.email or '',
        'section': s.section or '',
    } for s in students]

    return Response({
        'course': {'id': course.id, 'code': course.code, 'name': course.name, 'section': course.section},
        'students': student_list,
        'total': len(student_list)
    })

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def remove_student_from_course(request, course_id, student_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
        student = StudentProfile.objects.get(id=student_id, course=course)
        student.delete()
        return Response({'message': f'Student {student.matric_number} removed from course'})
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)
    except StudentProfile.DoesNotExist:
        return Response({'error': 'Student not found'}, status=404)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def clear_all_enrollments(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
        count = StudentProfile.objects.filter(course=course).count()
        StudentProfile.objects.filter(course=course).delete()
        return Response({'message': f'All enrollments cleared. {count} students removed.', 'removed_count': count})
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)




from django.db import transaction
# == COURSE MANAGEMENT ==
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_course_manual(request):
    """Manually create a course"""
    try:
        code = request.data.get('code')
        name = request.data.get('name')
        section = request.data.get('section', '')
        
        if not code or not name:
            return Response({'error': 'Course code and name are required'}, status=400)
        
        # Check if course already exists for this lecturer
        if Course.objects.filter(code=code, lecturer=request.user).exists():
            return Response({'error': f'Course {code} already exists'}, status=400)
        
        course = Course.objects.create(
            code=code,
            name=name,
            section=section,
            lecturer=request.user
        )
        ref = (
            Course.objects.filter(
                lecturer=request.user,
                semester_start__isnull=False,
            )
            .exclude(pk=course.pk)
            .first()
        )
        if ref:
            course.semester_start = ref.semester_start
            course.semester_end = ref.semester_end
            course.save(update_fields=['semester_start', 'semester_end'])

        return Response({
            'message': 'Course created successfully',
            'course': {
                'id': course.id,
                'code': course.code,
                'name': course.name,
                'section': course.section,
            }
        }, status=201)
        
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# --- SESSIONS ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def session_list(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    if request.method == 'GET':
        course_sessions = Session.objects.filter(course=course, is_finalized=False)
        return Response(SessionSerializer(course_sessions, many=True).data)

    if request.method == 'POST':
        serializer = SessionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(course=course)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def _past_session_payload(session):
    data = SessionSerializer(session).data
    data['course_code'] = session.course.code
    data['course_name'] = session.course.name
    return data


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def past_sessions_for_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    sessions = (
        Session.objects.filter(course=course, is_finalized=True)
        .select_related('course')
        .order_by('-date', '-start_time')
    )
    return Response([_past_session_payload(s) for s in sessions])


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def past_sessions_all(request):
    courses = Course.objects.filter(lecturer=request.user)
    sessions = (
        Session.objects.filter(course__in=courses, is_finalized=True)
        .select_related('course')
        .order_by('-date', '-start_time')
    )
    return Response([_past_session_payload(s) for s in sessions])


def _session_roster(session, request=None):
    enrolled = StudentProfile.objects.filter(course=session.course).order_by(
        'section', 'full_name'
    )
    records_by_matric = {
        r.matric_number: r
        for r in AttendanceRecord.objects.filter(session=session)
    }
    excuses_by_matric = {
        e.matric_number: e
        for e in AlertSessionExcuse.objects.filter(session=session)
    }
    roster = []
    for student in enrolled:
        record = records_by_matric.get(student.matric_number)
        excuse = excuses_by_matric.get(student.matric_number)
        is_pending = record and record.status == 'pending'
        roster.append({
            'id': record.id if record else None,
            'matric_number': student.matric_number,
            'full_name': student.full_name,
            'section': student.section or '',
            'status': record.status if record else 'absent',
            'scanned_at': record.scanned_at if record else None,
            'gps_verified': record.gps_verified if record else False,
            'latitude': record.latitude if record else None,
            'longitude': record.longitude if record else None,
            'is_pending': is_pending,
            'excuse': (
                AlertExcuseSerializer(excuse, context={'request': request}).data
                if excuse
                else None
            ),
        })

    # Sort to show pending attendances first
    roster.sort(key=lambda x: (not x['is_pending'], x['full_name']))
    return roster


# --- QR CODE GENERATION ---
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_qr(request, session_id):
    try:
        session = Session.objects.get(id=session_id)
    except Session.DoesNotExist:
        return Response({'error': 'Session not found'}, status=404)

    QRToken.objects.filter(session=session, is_active=True).update(is_active=False)

    token = uuid.uuid4()
    expires_at = timezone.now() + timedelta(seconds=180)  # 3 minutes validity
    QRToken.objects.create(
        session=session,
        token=token,
        expires_at=expires_at,
        is_active=True
    )

    # QR encodes a URL that students open on their phone
    frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:5173')
    scan_url = f"{frontend_url}/scan?token={token}"

    qr = qrcode.make(scan_url)
    buffer = io.BytesIO()
    qr.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    return Response({
        'token': str(token),
        'scan_url': scan_url,
        'expires_at': expires_at,
        'qr_image': f'data:image/png;base64,{qr_base64}'
    })


# --- MARK ATTENDANCE ---
@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def mark_attendance(request):
    token_value = request.data.get('token')
    full_name = request.data.get('full_name', '').strip()
    matric_number = request.data.get('matric_number', '').strip().upper()
    latitude = request.data.get('latitude')
    longitude = request.data.get('longitude')

    if not full_name or not matric_number:
        return Response({'error': 'Full name and matric number are required'}, status=400)

    try:
        qr_token = QRToken.objects.get(token=token_value, is_active=True)
    except QRToken.DoesNotExist:
        return Response({'error': 'Invalid or expired QR code'}, status=400)

    if timezone.now() > qr_token.expires_at:
        qr_token.is_active = False
        qr_token.save()
        return Response({'error': 'QR code has expired. Please scan the latest code.'}, status=400)

    session = qr_token.session
    course = session.course

    # Check if matric is enrolled in this course
    profile = StudentProfile.objects.filter(
        course=course,
        matric_number=matric_number
    ).first()

    if not profile:
        return Response({'error': f'Matric number {matric_number} is not enrolled in this course.'}, status=400)

    # GPS verification with geofence
    gps_verified = False
    attendance_status = 'pending'

    if latitude and longitude:
        try:
            lat = float(latitude)
            lon = float(longitude)

            # Check if course has geofence polygon
            if course.geofence_polygon:
                from alerts.geofencing import is_point_in_polygon
                gps_verified = is_point_in_polygon(lat, lon, course.geofence_polygon)
                # If inside geofence, mark as present; if outside, mark as pending
                attendance_status = 'present' if gps_verified else 'pending'
            else:
                # No geofence set, default to present
                gps_verified = True
                attendance_status = 'present'
        except (ValueError, TypeError):
            attendance_status = 'pending'

    # Check if already marked
    if AttendanceRecord.objects.filter(session=session, matric_number=matric_number).exists():
        return Response({'message': 'Attendance already marked!'}, status=200)

    AttendanceRecord.objects.create(
        session=session,
        full_name=profile.full_name,
        matric_number=matric_number,
        status=attendance_status,
        scanned_at=timezone.now(),
        latitude=float(latitude) if latitude else None,
        longitude=float(longitude) if longitude else None,
        gps_verified=gps_verified
    )

    if gps_verified:
        return Response({'message': f'Attendance marked successfully! ✅'}, status=201)
    elif attendance_status == 'pending':
        return Response({'message': f'Attendance flagged ⏳ — Lecturer will review your location.'}, status=201)
    else:
        return Response({'message': f'Attendance marked ⚠️ but GPS could not verify your location.'}, status=201)



# ---  live attendance view ---
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def session_attendance(request, session_id):
    try:
        session = Session.objects.select_related('course').get(id=session_id)
    except Session.DoesNotExist:
        return Response({'error': 'Session not found'}, status=404)

    if request.user.role == 'lecturer' and session.course.lecturer_id != request.user.id:
        return Response({'error': 'Not allowed'}, status=403)

    if session.is_finalized:
        return Response({
            'session': _past_session_payload(session),
            'roster': _session_roster(session, request),
        })

    records = AttendanceRecord.objects.filter(session=session).order_by('scanned_at')
    return Response(AttendanceRecordSerializer(records, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def session_excuse_student(request, session_id):
    """Upload proof and mark an absent student as excused (present for attendance)."""
    if request.user.role != 'lecturer':
        return Response({'error': 'Lecturers only'}, status=403)

    try:
        session = Session.objects.select_related('course').get(id=session_id)
    except Session.DoesNotExist:
        return Response({'error': 'Session not found'}, status=404)

    if session.course.lecturer_id != request.user.id:
        return Response({'error': 'Not allowed'}, status=403)

    if not session.is_finalized:
        return Response(
            {'error': 'Finalize this class before uploading excuses.'},
            status=400,
        )

    matric_number = (request.data.get('matric_number') or '').strip().upper()
    if not matric_number:
        return Response({'error': 'matric_number is required'}, status=400)

    reason_type = request.data.get('reason_type', '').strip()
    reason_note = request.data.get('reason_note', '')
    proof_file = request.FILES.get('proof')

    ok, err = excuse_session(
        session_id=session.id,
        matric_number=matric_number,
        course=session.course,
        proof_file=proof_file,
        reason_type=reason_type,
        reason_note=reason_note,
        alert=None,
        lecturer=request.user,
    )
    if not ok:
        return Response({'error': err}, status=400)

    return Response({
        'message': f'{matric_number} marked as excused for this class.',
        'roster': _session_roster(session, request),
    })

# --- TIMETABLE UPLOAD ---
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_timetable(request):
    if 'image' not in request.FILES:
        return Response({'error': 'No image uploaded'}, status=400)

    semester_start, err = _parse_semester_start(request.data.get('semester_start'))
    if err:
        return Response({'error': err}, status=400)

    image_file = request.FILES['image']
    image_bytes = image_file.read()

    try:
        sessions_data = parse_timetable_image(image_bytes, image_file.content_type)
    except Exception as e:
        return Response({'error': f'Failed to parse timetable: {str(e)}'}, status=500)

    day_map = {
        'Monday': 0, 'Tuesday': 1, 'Wednesday': 2,
        'Thursday': 3, 'Friday': 4
    }
    created_count = 0
    semester_end = semester_end_from_start(semester_start)

    for entry in sessions_data:
        try:
            display_name = entry.get('course_name') or entry['course_code']
            course, _ = Course.objects.get_or_create(
                code=entry['course_code'],
                defaults={'name': display_name, 'lecturer': request.user}
            )
            if course.name != display_name:
                course.name = display_name
                course.save(update_fields=['name'])
            if course.lecturer != request.user:
                course.lecturer = request.user
                course.save()

            course.semester_start = semester_start
            course.semester_end = semester_end
            course.save(update_fields=['semester_start', 'semester_end'])

            target_weekday = day_map.get(entry['day'], 0)
            days_ahead = (target_weekday - semester_start.weekday()) % 7
            first_session = semester_start + datetime.timedelta(days=days_ahead)

            for class_week in range(SEMESTER_TOTAL_CLASS_WEEKS):
                offset = calendar_week_offset(class_week)
                session_date = first_session + datetime.timedelta(weeks=offset)
                def _db_time(t):
                    t = (t or '').strip()
                    return t if len(t) > 5 else f'{t}:00'

                _, created_new = Session.objects.get_or_create(
                    course=course,
                    date=session_date,
                    start_time=_db_time(entry['start_time']),
                    end_time=_db_time(entry['end_time']),
                )
                if created_new:
                    created_count += 1
        except Exception as e:
            print(f"ERROR on entry {entry}: {e}")
            continue

    return Response({
        'message': (
            f'{created_count} sessions created (7 weeks, 1-week break, 7 weeks).'
        ),
        'semester_start': semester_start.isoformat(),
        'semester_end': semester_end.isoformat(),
    })

# --- DELETE COURSE ---
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
        course.delete()
        return Response({'message': 'Course deleted'})
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)


# --- DELETE SESSION ---
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_session(request, session_id):
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return Response({'message': 'Session deleted'})
    except Session.DoesNotExist:
        return Response({'error': 'Session not found'}, status=404)


def _time_to_seconds(t):
    return t.hour * 3600 + t.minute * 60 + t.second


def _session_not_ended(session, now=None):
    """True if the session has not ended yet (handles end before start as next day)."""
    now = timezone.localtime(now or timezone.now())
    start_secs = _time_to_seconds(session.start_time)
    end_secs = _time_to_seconds(session.end_time)
    now_secs = _time_to_seconds(now.time())
    if end_secs <= start_secs:
        # Overnight session: active from start until midnight, then until end.
        return now_secs >= start_secs or now_secs < end_secs
    return now_secs < end_secs


# --- TODAY'S SESSIONS ---
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def todays_sessions(request):
    """All unfinalized sessions scheduled for today (including in-progress and ended)."""
    today = timezone.localdate()
    courses = Course.objects.filter(lecturer=request.user)
    sessions = (
        Session.objects.filter(
            course__in=courses,
            date=today,
            is_finalized=False,
        )
        .select_related('course')
        .order_by('start_time')
    )
    data = SessionSerializer(sessions, many=True).data
    for item, session in zip(data, sessions):
        item['course_code'] = session.course.code
        item['is_active'] = _session_not_ended(session)
    return Response(data)

# --- SEMESTER DATES (no timetable required) ---
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_semester(request):
    """Set semester start/end on lecturer course(s) for bar-letter attendance window."""
    if request.user.role != 'lecturer':
        return Response({'error': 'Unauthorized'}, status=403)

    semester_start, err = _parse_semester_start(request.data.get('semester_start'))
    if err:
        return Response({'error': err}, status=400)

    semester_end = semester_end_from_start(semester_start)
    courses = Course.objects.filter(lecturer=request.user)
    course_id = request.data.get('course_id')
    if course_id is not None:
        courses = courses.filter(id=course_id)
        if not courses.exists():
            return Response({'error': 'Course not found'}, status=404)

    updated = courses.update(
        semester_start=semester_start,
        semester_end=semester_end,
    )

    if updated == 0:
        return Response({
            'message': (
                'Semester dates saved. They will apply when you add or import courses.'
            ),
            'semester_start': semester_start.isoformat(),
            'semester_end': semester_end.isoformat(),
            'updated_count': 0,
        })

    return Response({
        'message': f'Semester dates set on {updated} course(s).',
        'semester_start': semester_start.isoformat(),
        'semester_end': semester_end.isoformat(),
        'updated_count': updated,
    })


# --- EDIT COURSE ---
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def edit_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    name = request.data.get('name', course.name)
    code = request.data.get('code', course.code)
    course.name = name
    course.code = code

    if 'semester_start' in request.data:
        start, err = _parse_semester_start(request.data.get('semester_start'))
        if err:
            return Response({'error': err}, status=400)
        course.semester_start = start
        course.semester_end = semester_end_from_start(start)

    course.save()
    return Response(CourseSerializer(course).data)

# --- EXPORT ATTENDANCE TO EXCEL ---
from rest_framework.authtoken.models import Token
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def export_attendance_excel(request, session_id):
    # Authenticate via query param token
    token_key = request.GET.get('token')
    try:
        token = Token.objects.get(key=token_key)
        request.user = token.user
    except Token.DoesNotExist:
        return HttpResponse('Unauthorized', status=401)

    try:
        session = Session.objects.get(id=session_id)
    except Session.DoesNotExist:
        return HttpResponse('Session not found', status=404)

    records = AttendanceRecord.objects.filter(
        session=session
    ).order_by('scanned_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Attendance Report — {session.course.code} | {session.date} | {session.start_time} - {session.end_time}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center')

    headers = ['#', 'Full Name', 'Matric Number', 'Time Scanned', 'GPS Verified', 'Status']
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for idx, record in enumerate(records, start=1):
        scanned_at = record.scanned_at.strftime('%I:%M:%S %p') if record.scanned_at else '-'
        gps = 'Verified ✓' if record.gps_verified else 'Unverified'
        row = [idx, record.full_name, record.matric_number, scanned_at, gps, record.status.capitalize()]

        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 2, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

    summary_row = len(records) + 4
    ws.cell(row=summary_row, column=1, value='Total Present:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(records)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_{session.course.code}_{session.date}.xlsx'
    wb.save(response)
    return response

from .student_list_parser import parse_student_list
from .models import Course, Session, QRToken, AttendanceRecord, StudentProfile
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_student_list(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    if 'file' not in request.FILES:
        return Response({'error': 'No file uploaded'}, status=400)

    file_bytes = request.FILES['file'].read()

    try:
        data = parse_student_list(file_bytes, request.FILES['file'].content_type)
    except Exception as e:
        return Response({'error': f'Failed to parse: {str(e)}'}, status=500)

    total_added = 0
    sections_found = []

    for entry in data:
        for student in entry['students']:
            _, created = StudentProfile.objects.update_or_create(
                course=course,
                matric_number=student['matric_number'],
                defaults={
                    'full_name': student['full_name'],
                    'phone': student['phone'],
                    'email': student['email'],
                    'section': entry['section'],
                }
            )
            if created:
                total_added += 1
        sections_found.append(f"Seksyen {entry['section']} — {len(entry['students'])} students")

    return Response({
        'message': f'{total_added} students enrolled in {course.code}.',
        'sections': sections_found
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_list(request, course_id):
    try:
        course = Course.objects.get(id=course_id, lecturer=request.user)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found'}, status=404)

    students = StudentProfile.objects.filter(course=course).order_by('section', 'full_name')
    data = [
        {
            'id': s.id,
            'matric_number': s.matric_number,
            'full_name': s.full_name,
            'section': s.section or '',
            'email': s.email or '',
            'phone': s.phone or '',
        }
        for s in students
    ]
    return Response({'course': course.code, 'total': len(data), 'students': data})



@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def reset_semester(request):
    if not request.user.role == 'lecturer':
        return Response({'error': 'Unauthorized'}, status=403)

    courses = Course.objects.filter(lecturer=request.user)
    course_count = courses.count()

    # Delete everything cascade
    courses.delete()

    return Response({
        'message': f'Semester reset complete. {course_count} courses and all related sessions, attendance records and student profiles deleted.'
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finalize_session(request, session_id):
    try:
        session = Session.objects.select_related('course').get(id=session_id)
    except Session.DoesNotExist:
        return Response({'error': 'Session not found'}, status=404)

    if request.user.role == 'lecturer' and session.course.lecturer_id != request.user.id:
        return Response({'error': 'Not allowed'}, status=403)

    if session.is_finalized:
        return Response({'error': 'Session is already finalized.'}, status=400)

    course = session.course

    enrolled = StudentProfile.objects.filter(course=course)
    if not enrolled.exists():
        return Response({'error': 'No students enrolled in this course yet.'}, status=400)

    QRToken.objects.filter(session=session, is_active=True).update(is_active=False)

    present_matrics = set(
        AttendanceRecord.objects.filter(
            session=session, status='present'
        ).values_list('matric_number', flat=True)
    )

    absent_count = 0
    present_count = len(present_matrics)

    for student in enrolled:
        if student.matric_number not in present_matrics:
            _, created = AttendanceRecord.objects.get_or_create(
                session=session,
                matric_number=student.matric_number,
                defaults={
                    'full_name': student.full_name,
                    'status': 'absent',
                    'scanned_at': None,
                    'gps_verified': False,
                }
            )
            if created:
                absent_count += 1

    alerts_triggered = []
    for student in enrolled:
        before = Alert.objects.filter(
            course=course,
            matric_number=student.matric_number,
            is_sent=False,
        ).count()

        check_and_trigger_alerts(student.matric_number, course)

        after = Alert.objects.filter(
            course=course,
            matric_number=student.matric_number,
            is_sent=False,
        ).count()

        if after > before:
            alerts_triggered.append(student.matric_number)

    session.is_finalized = True
    session.save()

    return Response({
        'message': 'Session finalized.',
        'present': present_count,
        'absent': absent_count,
        'total_enrolled': enrolled.count(),
        'alerts_pending_review': len(alerts_triggered),
        'alerted_students': alerts_triggered,
    })


# --- APPROVE/REJECT PENDING ATTENDANCE ---
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_pending_attendance(request):
    """Lecturer approves flagged attendance (marks as present)."""
    attendance_id = request.data.get('attendance_id')
    try:
        record = AttendanceRecord.objects.get(id=attendance_id)
    except AttendanceRecord.DoesNotExist:
        return Response({'error': 'Attendance record not found'}, status=404)

    if record.session.course.lecturer != request.user:
        return Response({'error': 'Only the course lecturer can approve attendance'}, status=403)

    if record.status != 'pending':
        return Response({'error': 'Only pending attendance can be approved'}, status=400)

    record.status = 'present'
    record.save(update_fields=['status'])

    return Response({'message': f'Attendance approved for {record.matric_number}'}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_pending_attendance(request):
    """Lecturer rejects flagged attendance (marks as absent)."""
    attendance_id = request.data.get('attendance_id')
    try:
        record = AttendanceRecord.objects.get(id=attendance_id)
    except AttendanceRecord.DoesNotExist:
        return Response({'error': 'Attendance record not found'}, status=404)

    if record.session.course.lecturer != request.user:
        return Response({'error': 'Only the course lecturer can reject attendance'}, status=403)

    if record.status != 'pending':
        return Response({'error': 'Only pending attendance can be rejected'}, status=400)

    record.status = 'absent'
    record.save(update_fields=['status'])

    return Response({'message': f'Attendance rejected for {record.matric_number}'}, status=200)